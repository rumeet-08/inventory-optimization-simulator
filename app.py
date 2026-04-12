
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import math
import io
from groq import Groq
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

GROQ_API_KEY = st.secrets["GROQ_API_KEY"]

st.set_page_config(
    page_title="StockSense — Inventory Optimizer",
    page_icon="📦",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');

html, body, [class*="css"], .stApp {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    background-color: #111827 !important;
    color: #e5e7eb !important;
}

#MainMenu, footer, header { visibility: hidden; }

.block-container {
    padding: 2rem 2rem 2rem 2rem !important;
    max-width: 1200px !important;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: #1f2937 !important;
    border-radius: 12px !important;
    padding: 4px !important;
    gap: 4px !important;
    border: 1px solid #374151 !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    color: #9ca3af !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    padding: 8px 20px !important;
    border: none !important;
}
.stTabs [aria-selected="true"] {
    background: #374151 !important;
    color: #f9fafb !important;
}

/* Buttons */
.stButton > button {
    background: linear-gradient(135deg, #4f46e5, #7c3aed) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.65rem 1.5rem !important;
    font-weight: 600 !important;
    font-size: 13px !important;
    width: 100% !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
}

.stDownloadButton > button {
    background: #1f2937 !important;
    color: #818cf8 !important;
    border: 1px solid #4f46e5 !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    width: 100% !important;
}

/* Selectbox */
.stSelectbox > div > div {
    background: #1f2937 !important;
    border: 1px solid #374151 !important;
    border-radius: 10px !important;
    color: #f9fafb !important;
}

/* File uploader */
.stFileUploader > div {
    background: #1f2937 !important;
    border: 2px dashed #374151 !important;
    border-radius: 12px !important;
}

/* Chat */
.stChatMessage {
    background: #1f2937 !important;
    border: 1px solid #374151 !important;
    border-radius: 12px !important;
}

/* Dataframe */
.stDataFrame {
    background: #1f2937 !important;
}

/* Spinner */
.stSpinner > div {
    border-top-color: #6366f1 !important;
}

/* Custom components */
.ss-header {
    background: #1f2937;
    border: 1px solid #374151;
    border-radius: 16px;
    padding: 20px 24px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.ss-logo { display: flex; align-items: center; gap: 10px; }
.ss-logo-icon {
    width: 36px; height: 36px;
    background: linear-gradient(135deg, #4f46e5, #7c3aed);
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 18px;
}
.ss-logo-text { font-size: 18px; font-weight: 700; color: #f9fafb; letter-spacing: -0.02em; }
.ss-badge {
    font-size: 10px; background: #1e1b4b; color: #818cf8;
    padding: 3px 10px; border-radius: 20px; border: 1px solid #4f46e544;
}
.ss-status { font-size: 11px; color: #6b7280; display: flex; align-items: center; gap: 6px; }
.ss-dot { width: 6px; height: 6px; border-radius: 50%; background: #22c55e; display: inline-block; box-shadow: 0 0 6px #22c55e88; }

.ss-hero {
    background: linear-gradient(135deg, #1e1b4b 0%, #1f2937 100%);
    border: 1px solid #374151;
    border-radius: 16px;
    padding: 28px 32px;
    margin-bottom: 20px;
    position: relative;
    overflow: hidden;
}
.ss-hero-tag {
    font-size: 10px; font-weight: 600; letter-spacing: 0.12em;
    text-transform: uppercase; color: #818cf8; margin-bottom: 10px;
}
.ss-hero-title {
    font-size: 26px; font-weight: 700; color: #f9fafb;
    letter-spacing: -0.02em; line-height: 1.2; margin-bottom: 8px;
}
.ss-hero-accent {
    background: linear-gradient(90deg, #818cf8, #a78bfa);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
.ss-hero-sub { font-size: 13px; color: #9ca3af; line-height: 1.7; max-width: 540px; margin-bottom: 16px; }
.ss-pills { display: flex; flex-wrap: wrap; gap: 6px; }
.ss-pill {
    font-size: 10px; padding: 4px 12px; border-radius: 20px;
    border: 1px solid #374151; color: #6b7280; background: #111827;
}

.ss-card {
    background: #1f2937;
    border: 1px solid #374151;
    border-radius: 14px;
    padding: 20px 22px;
    margin-bottom: 12px;
}
.ss-card-title { font-size: 14px; font-weight: 600; color: #f9fafb; margin-bottom: 6px; }
.ss-card-desc { font-size: 12px; color: #9ca3af; line-height: 1.65; margin-bottom: 14px; }
.ss-step-badge {
    width: 28px; height: 28px; border-radius: 8px;
    background: linear-gradient(135deg, #4f46e5, #7c3aed);
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 12px; font-weight: 700; color: white; margin-bottom: 12px;
}
.ss-hint {
    background: #111827; border: 1px solid #374151;
    border-radius: 8px; padding: 8px 12px;
    font-size: 11px; margin-bottom: 14px;
    display: flex; align-items: center; gap: 6px;
}

.ss-metric {
    background: #1f2937; border: 1px solid #374151;
    border-radius: 12px; padding: 16px 18px; margin-bottom: 8px;
}
.ss-metric-label { font-size: 10px; color: #6b7280; text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 4px; }
.ss-metric-val { font-size: 22px; font-weight: 700; color: #f9fafb; }
.ss-metric-unit { font-size: 11px; color: #6b7280; margin-top: 2px; }

.ss-winner {
    background: linear-gradient(135deg, #1e1b4b, #2e1065);
    border: 1px solid #4f46e5; border-radius: 12px;
    padding: 14px 20px; margin-bottom: 16px;
    display: flex; align-items: center; justify-content: space-between;
    flex-wrap: wrap; gap: 12px;
}
.ss-winner-label { font-size: 10px; color: #a5b4fc; font-weight: 500; }
.ss-winner-val { font-size: 15px; font-weight: 700; color: #e0e7ff; }

.ss-risk-high { background: #450a0a; border: 1px solid #7f1d1d; color: #fca5a5; border-radius: 8px; padding: 8px 14px; font-size: 12px; margin-bottom: 12px; }
.ss-risk-med  { background: #451a03; border: 1px solid #78350f; color: #fcd34d; border-radius: 8px; padding: 8px 14px; font-size: 12px; margin-bottom: 12px; }
.ss-risk-low  { background: #052e16; border: 1px solid #14532d; color: #6ee7b7; border-radius: 8px; padding: 8px 14px; font-size: 12px; margin-bottom: 12px; }

.ss-scenario {
    background: #1f2937; border: 1px solid #374151;
    border-radius: 10px; padding: 12px 14px; margin-bottom: 8px;
}
.ss-scenario.active { border-color: #4f46e5; background: #1e1b4b; }
.ss-scenario-name { font-size: 13px; font-weight: 600; color: #e5e7eb; }
.ss-scenario-sl { font-size: 10px; color: #6b7280; margin-top: 2px; }

.ss-privacy {
    background: #1f2937; border: 1px solid #374151;
    border-radius: 10px; padding: 10px 14px;
    font-size: 11px; color: #6b7280; line-height: 1.6; margin-bottom: 16px;
}

.ss-info-grid {
    display: grid; grid-template-columns: repeat(4, 1fr);
    gap: 10px; margin-top: 20px;
}
.ss-info-card {
    background: #1f2937; border: 1px solid #374151;
    border-radius: 10px; padding: 14px;
    display: flex; gap: 10px;
}
.ss-info-icon { font-size: 16px; flex-shrink: 0; }
.ss-info-title { font-size: 11px; font-weight: 600; color: #9ca3af; margin-bottom: 3px; }
.ss-info-text { font-size: 10px; color: #6b7280; line-height: 1.55; }

.ss-section {
    font-size: 11px; font-weight: 600; letter-spacing: 0.08em;
    text-transform: uppercase; color: #6b7280;
    margin: 20px 0 12px; padding-bottom: 8px;
    border-bottom: 1px solid #1f2937;
}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════
# CORE MATH
# ════════════════════════════════════════════
Z_SCORES = {90: 1.28, 95: 1.645, 97: 1.88, 99: 2.326}

def calc_eoq(annual_demand, order_cost, holding_cost_per_unit):
    if holding_cost_per_unit <= 0 or annual_demand <= 0: return 0
    return math.sqrt((2 * annual_demand * order_cost) / holding_cost_per_unit)

def calc_safety_stock(z, demand_std, lead_time_avg, lead_time_std, daily_demand):
    return math.sqrt((z**2)*lead_time_avg*(demand_std**2) + (z**2)*(daily_demand**2)*(lead_time_std**2))

def calc_supplier_risk(num_suppliers, reliability):
    if num_suppliers == 1:   return 1 + ((100-reliability)/100)*1.5
    elif num_suppliers == 2: return 1 + ((100-reliability)/100)*0.8
    else:                    return 1 + ((100-reliability)/100)*0.4

def run_sku(row):
    results = {}
    monthly_demand = float(row["Monthly_Demand"])
    unit_cost      = float(row["Unit_Cost_USD"])
    order_cost     = float(row["Order_Cost_USD"])
    lead_time_avg  = float(row["Lead_Time_Days"])
    num_suppliers  = int(row["Num_Suppliers"])
    reliability    = float(row["Supplier_Reliability_Pct"])
    service_level  = float(row["Service_Level_Pct"])

    def safe_get(key, default):
        val = row.get(key, default)
        return default if pd.isna(val) else float(val)

    working_days   = safe_get("Working_Days_Month", 22)
    hold_pct       = safe_get("Monthly_Holding_Cost_Pct", 2.0)
    daily_demand   = monthly_demand / working_days
    demand_std     = safe_get("Demand_Std_Dev", daily_demand * 0.2)
    lt_std         = safe_get("Lead_Time_Std_Dev", lead_time_avg * 0.2)
    moq            = safe_get("MOQ", 0)
    shelf_life     = safe_get("Shelf_Life_Days", 9999)
    dead_stock     = safe_get("Dead_Stock_Units", 0)
    peak_mult      = safe_get("Peak_Season_Multiplier", 1.0)
    price_trend    = str(row.get("Price_Trend", "Stable")) if not pd.isna(row.get("Price_Trend", "Stable")) else "Stable"

    annual_demand  = monthly_demand * 12 * peak_mult
    annual_hold    = hold_pct * 12
    hold_per_unit  = unit_cost * (annual_hold / 100)
    risk_factor    = calc_supplier_risk(num_suppliers, reliability)
    price_adj      = 1.075 if price_trend=="Rising" else (0.93 if price_trend=="Falling" else 1.0)

    for name, sl in [("Conservative",99),("Balanced",95),("Lean",90)]:
        z   = Z_SCORES[sl]
        eoq = max(calc_eoq(annual_demand, order_cost, hold_per_unit) * price_adj, moq)
        if shelf_life < 9999: eoq = min(eoq, shelf_life * daily_demand * 0.8)
        ss  = calc_safety_stock(z, demand_std, lead_time_avg, lt_std, daily_demand) * risk_factor
        rop = daily_demand * lead_time_avg + ss
        hc  = ((eoq/2) + ss) * hold_per_unit
        oc  = (annual_demand/eoq)*order_cost if eoq>0 else 0
        wc  = dead_stock * unit_cost * (annual_hold/100)
        results[name] = {
            "eoq":round(eoq), "safety_stock":round(ss), "reorder_point":round(rop),
            "holding_cost":round(hc,2), "order_cost_ann":round(oc,2), "wc_cost":round(wc,2),
            "total_cost":round(hc+oc+wc,2), "orders_per_year":round(annual_demand/eoq if eoq>0 else 0,1),
            "stockout_risk":100-sl, "service_level":sl,
        }

    best = "Conservative" if service_level>=97 else ("Lean" if service_level<93 else "Balanced")
    if num_suppliers==1 and reliability<80: best="Conservative"
    if dead_stock > monthly_demand*2: best="Lean"

    results["recommended"] = best
    results["daily_demand"] = round(daily_demand,1)
    results["annual_demand"] = round(annual_demand)
    results["risk_score"] = round(((100-reliability)*0.4) + ((30 if num_suppliers==1 else 0)) + (dead_stock/(monthly_demand+1))*10, 1)
    return results

# ════════════════════════════════════════════
# CHARTS
# ════════════════════════════════════════════
C = {"Conservative":"#4f46e5","Balanced":"#7c3aed","Lean":"#a78bfa"}
BG = "#1f2937"
GRID = "#374151"
TEXT = "#9ca3af"

def base(title, h=300):
    return dict(
        title=dict(text=title, font=dict(color="#f9fafb",size=13), x=0),
        plot_bgcolor=BG, paper_bgcolor=BG,
        font=dict(color=TEXT, size=11),
        margin=dict(l=10,r=10,t=40,b=10), height=h,
        legend=dict(orientation="h",y=1.12,bgcolor="rgba(0,0,0,0)",font=dict(size=10)),
        xaxis=dict(showgrid=False, tickfont=dict(color=TEXT)),
        yaxis=dict(gridcolor=GRID, gridwidth=0.5, tickfont=dict(color=TEXT)),
        hoverlabel=dict(bgcolor="#374151",font_color="#f9fafb")
    )

def chart_costs(r):
    cats = ["Holding","Order","Dead Stock"]
    fig = go.Figure()
    for s in ["Conservative","Balanced","Lean"]:
        fig.add_trace(go.Bar(
            name=s, x=cats,
            y=[r[s]["holding_cost"],r[s]["order_cost_ann"],r[s]["wc_cost"]],
            marker_color=C[s],
            text=[f"${v:,.0f}" for v in [r[s]["holding_cost"],r[s]["order_cost_ann"],r[s]["wc_cost"]]],
            textposition="auto", textfont=dict(color="white",size=10)
        ))
    fig.update_layout(**base("Cost Breakdown ($)"), barmode="group")
    return fig

def chart_ss_rop(r):
    fig = make_subplots(rows=1,cols=2,subplot_titles=["Safety Stock","Reorder Point"],horizontal_spacing=0.15)
    for s in ["Conservative","Balanced","Lean"]:
        fig.add_trace(go.Bar(name=s,x=[s],y=[r[s]["safety_stock"]],marker_color=C[s],
            text=[f"{r[s]['safety_stock']:,}"],textposition="auto",textfont=dict(color="white",size=10),showlegend=False),row=1,col=1)
        fig.add_trace(go.Bar(name=s,x=[s],y=[r[s]["reorder_point"]],marker_color=C[s],
            text=[f"{r[s]['reorder_point']:,}"],textposition="auto",textfont=dict(color="white",size=10),showlegend=False),row=1,col=2)
    layout = base("Safety Stock & Reorder Point")
    fig.update_layout(**layout)
    fig.update_xaxes(showgrid=False,tickfont=dict(color=TEXT))
    fig.update_yaxes(gridcolor=GRID,gridwidth=0.5,tickfont=dict(color=TEXT))
    for ann in fig.layout.annotations: ann.font.color=TEXT; ann.font.size=11
    return fig

def chart_total(r):
    fig = go.Figure(go.Bar(
        x=["Conservative","Balanced","Lean"],
        y=[r[s]["total_cost"] for s in ["Conservative","Balanced","Lean"]],
        marker_color=[C[s] for s in ["Conservative","Balanced","Lean"]],
        text=[f"${r[s]['total_cost']:,.0f}" for s in ["Conservative","Balanced","Lean"]],
        textposition="auto", textfont=dict(color="white",size=11)
    ))
    fig.update_layout(**base("Total Annual Cost ($)"))
    return fig

def chart_risk(r):
    fig = go.Figure()
    for s in ["Conservative","Balanced","Lean"]:
        fig.add_trace(go.Scatter(
            x=[r[s]["stockout_risk"]],y=[r[s]["total_cost"]],
            mode="markers+text",
            marker=dict(size=20,color=C[s],line=dict(width=2,color="white")),
            text=[s],textposition="top center",textfont=dict(size=10,color=TEXT),name=s
        ))
    layout = base("Risk vs Cost Tradeoff")
    layout["xaxis"]["autorange"]="reversed"
    layout["xaxis"]["showgrid"]=True
    layout["xaxis"]["gridcolor"]=GRID
    layout["xaxis"]["title"]=dict(text="Stockout Risk (%)",font=dict(color=TEXT,size=10))
    layout["yaxis"]["title"]=dict(text="Annual Cost ($)",font=dict(color=TEXT,size=10))
    fig.update_layout(**layout)
    return fig

# ════════════════════════════════════════════
# AI
# ════════════════════════════════════════════
def get_ai_analysis(all_results, df):
    try:
        client = Groq(api_key=GROQ_API_KEY)
        top10  = sorted(all_results.items(), key=lambda x: x[1]["risk_score"], reverse=True)[:10]
        summary = ""
        for sku_id, r in top10:
            row = df[df["SKU_ID"]==sku_id].iloc[0]
            rec = r["recommended"]
            summary += f"\n- {sku_id} ({row['Product_Name']}): Risk={r['risk_score']}, {rec}, EOQ={r[rec]['eoq']}, SS={r[rec]['safety_stock']}, Cost=${r[rec]['total_cost']:,.0f}, Suppliers={row['Num_Suppliers']}, Reliability={row['Supplier_Reliability_Pct']}%"

        total_cost = sum(r[r["recommended"]]["total_cost"] for r in all_results.values())
        high_risk  = sum(1 for r in all_results.values() if r["risk_score"]>40)
        single_src = int((df["Num_Suppliers"]==1).sum())

        prompt = f"""Senior supply chain analyst. Portfolio: {len(all_results)} SKUs, ${total_cost:,.0f} total cost, {high_risk} high risk, {single_src} single source.
Top risks: {summary}

Respond with these headers:
**PORTFOLIO HEALTH SUMMARY**
**TOP 5 CRITICAL SKUs**
**3 BIGGEST RISKS**
**5 IMMEDIATE ACTIONS**
**COST SAVINGS OPPORTUNITIES**

Be specific with numbers."""

        r = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role":"user","content":prompt}],
            temperature=0.3, max_tokens=1200
        )
        return r.choices[0].message.content
    except Exception as e:
        return f"AI Error: {str(e)}"

def get_chat_response(client, question, history, all_results, df):
    try:
        ctx = f"Portfolio: {len(all_results)} SKUs\n"
        for sku_id, r in list(all_results.items())[:30]:
            row = df[df["SKU_ID"]==sku_id].iloc[0]
            rec = r["recommended"]
            ctx += f"- {sku_id} ({row['Product_Name']}): EOQ={r[rec]['eoq']}, SS={r[rec]['safety_stock']}, ROP={r[rec]['reorder_point']}, Cost=${r[rec]['total_cost']:,.0f}, Risk={r['risk_score']}\n"
        msgs = [{"role":"system","content":f"Senior inventory analyst. Data:\n{ctx}"}]
        for m in history: msgs.append(m)
        msgs.append({"role":"user","content":question})
        r = client.chat.completions.create(model="llama-3.3-70b-versatile",messages=msgs,temperature=0.3,max_tokens=600)
        return r.choices[0].message.content
    except Exception as e:
        return f"Error: {str(e)}"

# ════════════════════════════════════════════
# EXCEL
# ════════════════════════════════════════════
def generate_template():
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Inventory Data"
    req_fill = PatternFill("solid", fgColor="1e1b4b")
    opt_fill = PatternFill("solid", fgColor="1c1917")
    req_font = Font(bold=True, color="818cf8", size=10)
    opt_font = Font(bold=True, color="d97706", size=10)
    center   = Alignment(horizontal="center", vertical="center")
    req_cols = ["SKU_ID","Product_Name","Monthly_Demand","Unit_Cost_USD","Order_Cost_USD","Lead_Time_Days","Num_Suppliers","Supplier_Reliability_Pct","Service_Level_Pct"]
    opt_cols = ["Working_Days_Month","Monthly_Holding_Cost_Pct","Demand_Std_Dev","Lead_Time_Std_Dev","MOQ","Shelf_Life_Days","Dead_Stock_Units","Price_Trend","Peak_Season_Multiplier"]
    all_cols = req_cols + opt_cols
    ws.append(all_cols)
    for i, col in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = req_fill if col in req_cols else opt_fill
        cell.font = req_font if col in req_cols else opt_font
        cell.alignment = center
    samples = [
        ["SKU-001","Product A",4200,5.00,200,14,1,78,95,22,2.0,15,3,100,1095,200,"Stable",1.0],
        ["SKU-002","Product B",3100,8.50,180,21,2,88,95,22,2.0,12,4,50,730,0,"Rising",1.3],
        ["SKU-003","Product C",890,45.00,350,30,1,72,99,22,2.5,5,5,20,180,50,"Stable",1.0],
        ["SKU-004","Product D",5600,3.20,150,18,3,92,90,22,1.8,20,4,200,1825,0,"Falling",1.0],
        ["SKU-005","Product E",2400,12.00,220,25,1,65,99,22,2.0,10,6,50,730,300,"Rising",1.5],
    ]
    for s in samples:
        ws.append(s)
    for col in ws.columns:
        mx = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx+4,28)
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "HOW TO FILL THE TEMPLATE"
    ws2["A1"].font = Font(bold=True, size=13, color="818cf8")
    instructions = [
        ("",""),("INDIGO columns","REQUIRED — must fill for every product"),
        ("ORANGE columns","OPTIONAL — leave blank for smart defaults"),("",""),
        ("SKU_ID","Unique product code"),("Product_Name","Product name"),
        ("Monthly_Demand","Units sold per month"),("Unit_Cost_USD","Cost per unit in USD"),
        ("Order_Cost_USD","Cost to place one order"),("Lead_Time_Days","Avg supplier delivery days"),
        ("Num_Suppliers","Number of suppliers (1,2,3...)"),
        ("Supplier_Reliability_Pct","On-time delivery rate 0-100"),
        ("Service_Level_Pct","Target: 90, 95, 97, or 99"),("",""),
        ("Working_Days_Month","Default: 22"),("Monthly_Holding_Cost_Pct","Default: 2%"),
        ("Demand_Std_Dev","Default: 20% of daily demand"),
        ("Lead_Time_Std_Dev","Default: 20% of lead time"),
        ("MOQ","Minimum order quantity"),("Shelf_Life_Days","Product expiry in days"),
        ("Dead_Stock_Units","Idle units on hand"),
        ("Price_Trend","Rising / Stable / Falling"),
        ("Peak_Season_Multiplier","e.g. 1.8 for 80% demand spike"),
        ("",""),("MAX SKUs","500 per upload"),
    ]
    for r in instructions:
        ws2.append(list(r))
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 50
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out

def generate_report(all_results, df):
    wb = openpyxl.Workbook()
    hf = PatternFill("solid", fgColor="1e1b4b")
    hfont = Font(bold=True, color="818cf8", size=10)
    hr = PatternFill("solid", fgColor="450a0a")
    hm = PatternFill("solid", fgColor="451a03")
    hl = PatternFill("solid", fgColor="052e16")
    center = Alignment(horizontal="center")

    def hdr(ws, n):
        for c in range(1,n+1):
            ws.cell(row=1,column=c).fill=hf
            ws.cell(row=1,column=c).font=hfont
            ws.cell(row=1,column=c).alignment=center

    def aw(ws):
        for col in ws.columns:
            mx=max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(mx+4,28)

    ws1=wb.active; ws1.title="Summary"
    h1=["SKU ID","Product","Risk Score","Recommended","EOQ","Safety Stock","Reorder Point","Total Cost/Year","Orders/Year"]
    ws1.append(h1); hdr(ws1,len(h1))
    for sku_id,r in all_results.items():
        row=df[df["SKU_ID"]==sku_id].iloc[0]; rec=r["recommended"]
        ws1.append([sku_id,row["Product_Name"],r["risk_score"],rec,r[rec]["eoq"],r[rec]["safety_stock"],r[rec]["reorder_point"],f"${r[rec]['total_cost']:,.0f}",r[rec]["orders_per_year"]])
        fill=hr if r["risk_score"]>40 else (hm if r["risk_score"]>20 else hl)
        for c in range(1,len(h1)+1): ws1.cell(row=ws1.max_row,column=c).fill=fill
    aw(ws1)

    ws2=wb.create_sheet("High Risk"); h2=["SKU ID","Product","Risk Score","Reason","Suppliers","Reliability","Action"]
    ws2.append(h2); hdr(ws2,len(h2))
    for sku_id,r in sorted(all_results.items(),key=lambda x:x[1]["risk_score"],reverse=True):
        if r["risk_score"]<=40: continue
        row=df[df["SKU_ID"]==sku_id].iloc[0]
        reasons=[]
        if int(row["Num_Suppliers"])==1: reasons.append("Single source")
        if float(row["Supplier_Reliability_Pct"])<80: reasons.append("Low reliability")
        ws2.append([sku_id,row["Product_Name"],r["risk_score"]," | ".join(reasons) or "Multiple",row["Num_Suppliers"],row["Supplier_Reliability_Pct"],f"Increase SS to {r['Conservative']['safety_stock']}"])
        for c in range(1,len(h2)+1): ws2.cell(row=ws2.max_row,column=c).fill=hr
    aw(ws2)

    ws3=wb.create_sheet("Cost Analysis"); h3=["SKU ID","Product","Scenario","Holding","Order","Dead Stock","Total"]
    ws3.append(h3); hdr(ws3,len(h3))
    for sku_id,r in all_results.items():
        row=df[df["SKU_ID"]==sku_id].iloc[0]
        for sc in ["Conservative","Balanced","Lean"]:
            s=r[sc]; ws3.append([sku_id,row["Product_Name"],sc,f"${s['holding_cost']:,.0f}",f"${s['order_cost_ann']:,.0f}",f"${s['wc_cost']:,.0f}",f"${s['total_cost']:,.0f}"])
    aw(ws3)

    ws4=wb.create_sheet("Reorder Schedule"); h4=["SKU ID","Product","Daily Demand","Reorder Point","EOQ","Orders/Year","Lead Time","Safety Stock"]
    ws4.append(h4); hdr(ws4,len(h4))
    for sku_id,r in all_results.items():
        row=df[df["SKU_ID"]==sku_id].iloc[0]; rec=r["recommended"]
        ws4.append([sku_id,row["Product_Name"],r["daily_demand"],r[rec]["reorder_point"],r[rec]["eoq"],r[rec]["orders_per_year"],row["Lead_Time_Days"],r[rec]["safety_stock"]])
    aw(ws4)

    ws5=wb.create_sheet("Scenarios"); h5=["SKU ID","Product","Cons EOQ","Cons SS","Cons Cost","Bal EOQ","Bal SS","Bal Cost","Lean EOQ","Lean SS","Lean Cost","Recommended"]
    ws5.append(h5); hdr(ws5,len(h5))
    for sku_id,r in all_results.items():
        row=df[df["SKU_ID"]==sku_id].iloc[0]
        ws5.append([sku_id,row["Product_Name"],r["Conservative"]["eoq"],r["Conservative"]["safety_stock"],f"${r['Conservative']['total_cost']:,.0f}",r["Balanced"]["eoq"],r["Balanced"]["safety_stock"],f"${r['Balanced']['total_cost']:,.0f}",r["Lean"]["eoq"],r["Lean"]["safety_stock"],f"${r['Lean']['total_cost']:,.0f}",r["recommended"]])
    aw(ws5)

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ════════════════════════════════════════════
# APP HEADER
# ════════════════════════════════════════════
st.markdown("""
<div class="ss-header">
  <div class="ss-logo">
    <div class="ss-logo-icon">📦</div>
    <span class="ss-logo-text">StockSense</span>
    <span class="ss-badge">FREE</span>
  </div>
  <div class="ss-status">
    <span class="ss-dot"></span>
    Ready · No sign-up · Data never stored
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

# ════════════════════════════════════════════
# TABS
# ════════════════════════════════════════════
tab1, tab2, tab3 = st.tabs(["📂  Upload", "📊  Results", "🤖  AI & Export"])

# ══════════════════════════════
# TAB 1
# ══════════════════════════════
with tab1:
    st.markdown("""
    <div class="ss-hero">
      <div class="ss-hero-tag">Inventory Optimization Platform</div>
      <div class="ss-hero-title">Smart Inventory Decisions.<br><span class="ss-hero-accent">For Any Industry.</span></div>
      <div class="ss-hero-sub">Upload your product catalog and instantly get EOQ, safety stock, reorder points, and AI-powered scenario comparison — whether you manage 5 SKUs or 500.</div>
      <div class="ss-pills">
        <span class="ss-pill">🏭 Manufacturing</span>
        <span class="ss-pill">🛒 Retail</span>
        <span class="ss-pill">💊 Pharma</span>
        <span class="ss-pill">🍔 Food & Bev</span>
        <span class="ss-pill">💻 Electronics</span>
        <span class="ss-pill">🚚 Logistics</span>
        <span class="ss-pill">🏥 Healthcare</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown("""
        <div class="ss-card">
          <div class="ss-step-badge">1</div>
          <div class="ss-card-title">Download the Template</div>
          <div class="ss-card-desc">Get our Excel template and fill in your product details. Includes sample data and a full instructions sheet.</div>
          <div class="ss-hint">
            <span style="color:#818cf8;font-weight:600;">9 fields required</span>
            <span style="color:#6b7280;">· rest are optional with smart defaults</span>
          </div>
        </div>
        """, unsafe_allow_html=True)
        tmpl = generate_template()
        st.download_button("⬇️  Download Excel Template", tmpl,
                          "stocksense_template.xlsx",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          use_container_width=True)

    with col2:
        st.markdown("""
        <div class="ss-card">
          <div class="ss-step-badge">2</div>
          <div class="ss-card-title">Upload & Get Results</div>
          <div class="ss-card-desc">Upload your filled template. Results load automatically — just switch to the Results tab.</div>
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader("Upload Excel file", type=["xlsx"], label_visibility="collapsed")
        if uploaded:
            try:
                df = pd.read_excel(uploaded, sheet_name="Inventory Data")
                req = ["SKU_ID","Product_Name","Monthly_Demand","Unit_Cost_USD","Order_Cost_USD","Lead_Time_Days","Num_Suppliers","Supplier_Reliability_Pct","Service_Level_Pct"]
                missing = [c for c in req if c not in df.columns]
                if missing:
                    st.error(f"❌ Missing: {', '.join(missing)}")
                else:
                    df = df.dropna(subset=req).head(500)
                    if len(df)==0:
                        st.error("❌ No valid rows found.")
                    else:
                        st.success(f"✅ {len(df)} SKUs loaded — go to Results tab!")
                        st.session_state["df"] = df
                        if "all_results" in st.session_state:
                            del st.session_state["all_results"]
            except Exception as e:
                st.error(f"❌ {str(e)}")

    st.markdown("""
    <div class="ss-info-grid">
      <div class="ss-info-card"><div class="ss-info-icon">🔒</div><div><div class="ss-info-title">Data Privacy</div><div class="ss-info-text">Never stored. Deleted when you close the tab.</div></div></div>
      <div class="ss-info-card"><div class="ss-info-icon">⚡</div><div><div class="ss-info-title">Instant Results</div><div class="ss-info-text">All SKUs calculated across 3 scenarios in seconds.</div></div></div>
      <div class="ss-info-card"><div class="ss-info-icon">🤖</div><div><div class="ss-info-title">AI Analysis</div><div class="ss-info-text">Portfolio insights, risk flags, recommendations.</div></div></div>
      <div class="ss-info-card"><div class="ss-info-icon">📊</div><div><div class="ss-info-title">Excel Export</div><div class="ss-info-text">Full report with 5 organized sheets.</div></div></div>
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════
# TAB 2
# ══════════════════════════════
with tab2:
    if "df" not in st.session_state:
        st.markdown("<div style='text-align:center;padding:80px 0;color:#6b7280;font-size:14px;'>📂 Upload your data in the Upload tab first.</div>", unsafe_allow_html=True)
    else:
        df = st.session_state["df"]
        if "all_results" not in st.session_state:
            with st.spinner("Calculating all SKUs..."):
                all_results = {}
                for _, row in df.iterrows():
                    try: all_results[str(row["SKU_ID"])] = run_sku(row)
                    except: pass
                st.session_state["all_results"] = all_results
        all_results = st.session_state["all_results"]

        # Portfolio summary
        total_cost = sum(r[r["recommended"]]["total_cost"] for r in all_results.values())
        high_risk  = sum(1 for r in all_results.values() if r["risk_score"]>40)
        single_src = int((df["Num_Suppliers"]==1).sum())

        st.markdown("<div class='ss-section'>Portfolio Overview</div>", unsafe_allow_html=True)
        c1,c2,c3,c4 = st.columns(4)
        with c1: st.markdown(f'<div class="ss-metric"><div class="ss-metric-label">Total SKUs</div><div class="ss-metric-val" style="color:#818cf8">{len(all_results)}</div><div class="ss-metric-unit">analysed</div></div>', unsafe_allow_html=True)
        with c2: st.markdown(f'<div class="ss-metric"><div class="ss-metric-label">Total Annual Cost</div><div class="ss-metric-val">${total_cost:,.0f}</div><div class="ss-metric-unit">all SKUs</div></div>', unsafe_allow_html=True)
        with c3: st.markdown(f'<div class="ss-metric"><div class="ss-metric-label">High Risk SKUs</div><div class="ss-metric-val" style="color:#f87171">{high_risk}</div><div class="ss-metric-unit">need attention</div></div>', unsafe_allow_html=True)
        with c4: st.markdown(f'<div class="ss-metric"><div class="ss-metric-label">Single Source</div><div class="ss-metric-val" style="color:#fbbf24">{single_src}</div><div class="ss-metric-unit">supplier risk</div></div>', unsafe_allow_html=True)

        # SKU selector
        st.markdown("<div class='ss-section' style='margin-top:24px;'>SKU Analysis</div>", unsafe_allow_html=True)
        opts = [f"{row['SKU_ID']} — {row['Product_Name']}" for _,row in df.iterrows()]
        sel  = st.selectbox("Select SKU:", opts, label_visibility="collapsed")
        sid  = sel.split(" — ")[0]
        sname= sel.split(" — ")[1]
        r    = all_results[str(sid)]
        rec  = r["recommended"]

        # Winner banner
        st.markdown(f"""
        <div class="ss-winner">
          <div><div class="ss-winner-label">Recommended</div><div class="ss-winner-val">⭐ {rec}</div></div>
          <div><div class="ss-winner-label">Annual Cost</div><div class="ss-winner-val">${r[rec]['total_cost']:,.0f}</div></div>
          <div><div class="ss-winner-label">EOQ</div><div class="ss-winner-val">{r[rec]['eoq']:,} units</div></div>
          <div><div class="ss-winner-label">Safety Stock</div><div class="ss-winner-val">{r[rec]['safety_stock']:,} units</div></div>
          <div><div class="ss-winner-label">Reorder Point</div><div class="ss-winner-val">{r[rec]['reorder_point']:,} units</div></div>
        </div>
        """, unsafe_allow_html=True)

        rs = r["risk_score"]
        if rs>40:   st.markdown(f'<div class="ss-risk-high">🔴 High Risk (score: {rs}) — Immediate attention required</div>', unsafe_allow_html=True)
        elif rs>20: st.markdown(f'<div class="ss-risk-med">🟡 Medium Risk (score: {rs}) — Monitor closely</div>', unsafe_allow_html=True)
        else:       st.markdown(f'<div class="ss-risk-low">🟢 Low Risk (score: {rs}) — Well managed</div>', unsafe_allow_html=True)

        # Scenario columns
        s1,s2,s3 = st.columns(3)
        for col, sc in zip([s1,s2,s3],["Conservative","Balanced","Lean"]):
            s = r[sc]
            with col:
                active = "active" if sc==rec else ""
                st.markdown(f'<div class="ss-scenario {active}"><div class="ss-scenario-name">{"⭐ " if sc==rec else ""}{sc}</div><div class="ss-scenario-sl">SL: {s["service_level"]}% · Risk: {s["stockout_risk"]}%</div></div>', unsafe_allow_html=True)
                for label,val,unit in [("EOQ",f"{s['eoq']:,}","units/order"),("Safety Stock",f"{s['safety_stock']:,}","units"),("Reorder Point",f"{s['reorder_point']:,}","units"),("Total Cost",f"${s['total_cost']:,.0f}","USD/year"),("Orders/Year",str(s['orders_per_year']),"orders")]:
                    st.markdown(f'<div class="ss-metric"><div class="ss-metric-label">{label}</div><div class="ss-metric-val" style="font-size:18px;">{val}</div><div class="ss-metric-unit">{unit}</div></div>', unsafe_allow_html=True)

        # Charts
        st.markdown("<div class='ss-section' style='margin-top:24px;'>Visual Analysis</div>", unsafe_allow_html=True)
        ch1,ch2 = st.columns(2)
        with ch1: st.plotly_chart(chart_costs(r), use_container_width=True, config={"displayModeBar":False})
        with ch2: st.plotly_chart(chart_ss_rop(r), use_container_width=True, config={"displayModeBar":False})
        ch3,ch4 = st.columns(2)
        with ch3: st.plotly_chart(chart_total(r), use_container_width=True, config={"displayModeBar":False})
        with ch4: st.plotly_chart(chart_risk(r), use_container_width=True, config={"displayModeBar":False})

# ══════════════════════════════
# TAB 3
# ══════════════════════════════
with tab3:
    if "all_results" not in st.session_state:
        st.markdown("<div style='text-align:center;padding:80px 0;color:#6b7280;font-size:14px;'>📂 Upload your data first.</div>", unsafe_allow_html=True)
    else:
        all_results = st.session_state["all_results"]
        df          = st.session_state["df"]

        st.markdown('<div class="ss-privacy">🔒 <strong style="color:#9ca3af;">Privacy:</strong> Data processed locally, never stored. AI receives summarised metrics only — not your raw data. Everything deleted when you close this tab.</div>', unsafe_allow_html=True)

        st.markdown("<div class='ss-section'>AI Portfolio Analysis</div>", unsafe_allow_html=True)
        if st.button("🤖  Generate AI Analysis", use_container_width=True):
            with st.spinner("Analysing your portfolio..."):
                ai_out = get_ai_analysis(all_results, df)
                st.session_state["ai_output"]     = ai_out
                st.session_state["chat_messages"] = []
                st.session_state["chat_history"]  = []

        if "ai_output" in st.session_state:
            st.markdown(st.session_state["ai_output"])

            st.markdown("<div class='ss-section' style='margin-top:24px;'>Ask Follow-up Questions</div>", unsafe_allow_html=True)
            st.caption("Ask anything about your inventory — SKUs, reorder decisions, risk explanations, cost savings.")

            if "chat_messages" not in st.session_state: st.session_state["chat_messages"]=[]
            if "chat_history"  not in st.session_state: st.session_state["chat_history"] =[]

            for msg in st.session_state["chat_messages"]:
                with st.chat_message(msg["role"]): st.markdown(msg["content"])

            if prompt := st.chat_input("Ask a question..."):
                st.session_state["chat_messages"].append({"role":"user","content":prompt})
                with st.chat_message("user"): st.markdown(prompt)
                with st.chat_message("assistant"):
                    with st.spinner("Thinking..."):
                        client = Groq(api_key=GROQ_API_KEY)
                        resp   = get_chat_response(client, prompt, st.session_state["chat_history"], all_results, df)
                        st.markdown(resp)
                st.session_state["chat_messages"].append({"role":"assistant","content":resp})
                st.session_state["chat_history"].append({"role":"user","content":prompt})
                st.session_state["chat_history"].append({"role":"assistant","content":resp})

            if st.session_state.get("chat_messages"):
                if st.button("🗑️ Clear Chat", use_container_width=True):
                    st.session_state["chat_messages"]=[]
                    st.session_state["chat_history"] =[]
                    st.rerun()

        st.markdown("<div class='ss-section' style='margin-top:28px;'>Download Report</div>", unsafe_allow_html=True)
        st.caption("Full Excel report with 5 organized sheets — summary, high risk, cost analysis, reorder schedule, scenario comparison.")
        report = generate_report(all_results, df)
        st.download_button("⬇️  Download Full Excel Report", report,
                          "stocksense_report.xlsx",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          use_container_width=True)
